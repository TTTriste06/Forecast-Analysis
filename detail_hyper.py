import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO

def add_detail_link_and_sheets(wb, ws_main, df_order, df_sales, df_forecast, all_months):
    """
    为主计划中的每个非零数据单元格添加超链接，
    并生成对应的原始数据明细 Sheet（如“订单-品名-2025-07”）
    """
    created_sheets = set()
    
    for row in range(3, ws_main.max_row + 1):
        item_name = ws_main.cell(row=row, column=3).value  # 品名在第3列
        for i, ym in enumerate(all_months):
            for offset, source, df, date_col, value_col, prefix in [
                (0, "预测", df_forecast, None, f"{ym}-预测", "预测"),
                (1, "订单", df_order, "客户要求交期", "订单数量", "订单"),
                (2, "出货", df_sales, "交易日期", "数量", "出货"),
            ]:
                col = 4 + i * 3 + offset
                cell = ws_main.cell(row=row, column=col)
                val = cell.value
                if not val or float(val) == 0:
                    continue

                sheet_name = f"{prefix}-{item_name}-{ym}"
                sheet_name = sheet_name[:31]  # Excel 限制
                cell.hyperlink = HYPERLINK("#\'{sheet_name}\'!A1", "{cell.value}")'

                cell.font = Font(underline="single", color="0000FF")

                if sheet_name in created_sheets:
                    continue
                created_sheets.add(sheet_name)

                if prefix == "预测":
                    # 从预测中找出匹配“生产料号”==品名的行
                    month_pattern = f"{int(ym[-2:])}月预测"
                    df_filtered = df[df["生产料号"] == item_name][["生产料号", month_pattern]] if month_pattern in df.columns else pd.DataFrame()
                else:
                    df_temp = df.copy()
                    if date_col in df_temp.columns:
                        df_temp[date_col] = pd.to_datetime(df_temp[date_col], errors="coerce")
                        df_temp["年月"] = df_temp[date_col].dt.to_period("M").astype(str)
                        df_filtered = df_temp[(df_temp["品名"] == item_name) & (df_temp["年月"] == ym)]
                    else:
                        df_filtered = pd.DataFrame()

                ws_detail = wb.create_sheet(sheet_name)
                if not df_filtered.empty:
                    for r_idx, row_data in enumerate(dataframe_to_rows(df_filtered, index=False, header=True), start=1):
                        for c_idx, val in enumerate(row_data, start=1):
                            ws_detail.cell(row=r_idx, column=c_idx, value=val)
                else:
                    ws_detail.cell(row=1, column=1, value="无匹配数据")

    return wb
