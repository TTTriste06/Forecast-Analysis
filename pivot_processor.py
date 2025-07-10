import pandas as pd
import re
from io import BytesIO
import streamlit as st
from openpyxl.styles import Alignment, Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from urllib.parse import quote
from mapping_utils import (
    apply_mapping_and_merge, 
    apply_extended_substitute_mapping,
    split_mapping_data
)
from info_extract import (
    extract_all_year_months, 
    fill_forecast_data, 
    fill_order_data, 
    fill_sales_data, 
    highlight_by_detecting_column_headers
)
from choose_utils import extract_detail_from_click

class PivotProcessor:
    def process(self, template_df, forecast_df, order_df, sales_df, mapping_file):
        # Step 1: Apply mappings
        mapping_df = clean_mapping_headers(mapping_file)
        main_df = replace_all_names_with_mapping(template_df.copy(), mapping_df)

        # Step 2: Fill basic info
        fill_spec_and_wafer_info(main_df, mapping_df)
        fill_packaging_info(main_df, mapping_df)

        # Step 3: Append forecast
        append_forecast_to_summary(main_df, forecast_df)
        merge_forecast_header(main_df)

        # Step 4: Append unfulfilled
        append_unfulfilled_summary_columns_by_date(main_df, order_df)
        merge_unfulfilled_order_header(main_df)

        # Step 5: Append sales
        append_sales_summary_columns(main_df, sales_df)

        # Step 6: Generate plans
        forecast_months = sorted(set(re.findall(r"(\d{4}-\d{2})", "".join(main_df.columns))))
        generate_monthly_plan_fields(main_df, forecast_months)
        generate_monthly_semi_plan(main_df, forecast_months, mapping_df)
        generate_monthly_gap_columns(main_df, forecast_months)
        append_wafer_plan_columns(main_df, forecast_months)
        append_wafer_gap_columns(main_df, forecast_months)

        # Step 7: Format
        format_currency_columns_rmb(main_df)
        format_thousands_separator(main_df)

        # Output
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            main_df.to_excel(writer, index=False, sheet_name="主计划", startrow=1)
            wb = writer.book
            ws = writer.sheets["主计划"]
            ws.cell(row=1, column=1, value=f"主计划生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            legend_cell = ws.cell(row=1, column=3)
            legend_cell.value = "Red < 0    Yellow < 安全库存    Orange > 2 × 安全库存"
            legend_cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            legend_cell.fill = PatternFill(start_color="FFCCE6FF", end_color="FFCCE6FF", fill_type="solid")

            adjust_column_width(ws)
            highlight_replaced_names_in_main_sheet(ws, main_df)

        output.seek(0)
        return main_df, output
