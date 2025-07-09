import re
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def extract_all_year_months(df_forecast, df_order, df_sales):
    # 1. 从 forecast header 提取 x月预测 列中的月份
    month_pattern = re.compile(r"(\d{1,2})月预测")
    forecast_months = []
    for col in df_forecast.columns:
        match = month_pattern.match(str(col))
        if match:
            month = match.group(1).zfill(2)
            forecast_months.append(f"2025-{month}")  # ✅ 根据需要调整年份

    # 2. 从 order 文件第 B 列（假设是“订单日期”）
    order_date_col = df_order.columns[11]
    df_order[order_date_col] = pd.to_datetime(df_order[order_date_col], errors="coerce")
    order_months = (
        df_order[order_date_col]
        .dropna()
        .dt.to_period("M")
        .astype(str)
        .loc[lambda x: x != "NaT"]
        .unique()
        .tolist()
    )

    # 3. 从 sales 文件第 F 列（假设是“交易日期”）
    sales_date_col = df_sales.columns[5]
    df_sales[sales_date_col] = pd.to_datetime(df_sales[sales_date_col], errors="coerce")
    sales_months = (
        df_sales[sales_date_col]
        .dropna()
        .dt.to_period("M")
        .astype(str)
        .loc[lambda x: x != "NaT"]
        .unique()
        .tolist()
    )

    # 合并并去重
    all_months = sorted(set(forecast_months + order_months + sales_months))

    # 生成从最小到最大之间的所有月份
    if all_months:
        min_month = pd.Period(min(all_months), freq="M")
        max_month = pd.Period(max(all_months), freq="M")
        full_months = [str(p) for p in pd.period_range(min_month, max_month, freq="M")]
    else:
        full_months = []
    
    return full_months


def fill_forecast_data(main_df, df_forecast, forecast_months, source_map=None):
    df_forecast["生产料号"] = df_forecast["生产料号"].astype(str).str.strip()
    df_forecast["品名"] = df_forecast["生产料号"]

    month_pattern = re.compile(r"(\d{1,2})月预测")
    forecast_cols = {
        f"2025-{match.group(1).zfill(2)}": col
        for col in df_forecast.columns
        if (match := month_pattern.match(str(col)))
    }

    for ym in forecast_months:
        colname = f"{ym}-预测"
        if colname in main_df.columns and ym in forecast_cols:
            month_col = forecast_cols[ym]
            grouped = df_forecast.groupby("品名")[month_col].sum(min_count=1)
            main_df[colname] = main_df["品名"].map(grouped).fillna(0)

            if source_map is not None:
                for idx, row in df_forecast.iterrows():
                    name = row["品名"]
                    val = row.get(month_col, 0)
                    if pd.notna(val) and val != 0:
                        source_map[(name, ym, "预测")].append({
                            "来源": "forecast",
                            "来源行号": idx + 2,
                            "字段值": val
                        })
    return main_df

def fill_order_data(main_df, df_order, forecast_months, source_map=None):
    df_order = df_order.copy()
    df_order["客户要求交期"] = pd.to_datetime(df_order["客户要求交期"], errors="coerce")
    df_order["年月"] = df_order["客户要求交期"].dt.to_period("M").astype(str)
    df_order["订单数量"] = pd.to_numeric(df_order["订单数量"], errors="coerce").fillna(0)

    grouped = df_order.groupby(["品名", "年月"])["订单数量"].sum().unstack().fillna(0)

    for ym in forecast_months:
        colname = f"{ym}-订单"
        if colname in main_df.columns and ym in grouped.columns:
            main_df[colname] = main_df["品名"].map(grouped[ym]).fillna(0)

    if source_map is not None:
        for idx, row in df_order.iterrows():
            name = row["品名"]
            ym = row["年月"]
            val = row["订单数量"]
            if pd.notna(val) and val != 0 and ym in forecast_months:
                source_map[(name, ym, "订单")].append({
                    "来源": "order",
                    "来源行号": idx + 2,
                    "字段值": val
                })
    return main_df

def fill_sales_data(main_df, df_sales, forecast_months, source_map=None):
    df_sales = df_sales.copy()
    df_sales["交易日期"] = pd.to_datetime(df_sales["交易日期"], errors="coerce")
    df_sales["年月"] = df_sales["交易日期"].dt.to_period("M").astype(str)
    df_sales["数量"] = pd.to_numeric(df_sales["数量"], errors="coerce").fillna(0)

    grouped = df_sales.groupby(["品名", "年月"])["数量"].sum().unstack().fillna(0)

    for ym in forecast_months:
        colname = f"{ym}-出货"
        if colname in main_df.columns and ym in grouped.columns:
            main_df[colname] = main_df["品名"].map(grouped[ym]).fillna(0)

    if source_map is not None:
        for idx, row in df_sales.iterrows():
            name = row["品名"]
            ym = row["年月"]
            val = row["数量"]
            if pd.notna(val) and val != 0 and ym in forecast_months:
                source_map[(name, ym, "出货")].append({
                    "来源": "sales",
                    "来源行号": idx + 2,
                    "字段值": val
                })
    return main_df

def highlight_by_detecting_column_headers(ws):
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    header = [cell.value for cell in ws[2]]
    column_pairs = []
    for i in range(len(header) - 1):
        name1 = str(header[i]).strip()
        name2 = str(header[i + 1]).strip()
        if name1.endswith("预测") and name2.endswith("订单"):
            column_pairs.append((i + 1, i + 2))

    for row in range(3, ws.max_row + 1):
        for forecast_col, order_col in column_pairs:
            cell_forecast = ws.cell(row=row, column=forecast_col)
            cell_order = ws.cell(row=row, column=order_col)
            try:
                val_forecast = float(cell_forecast.value or 0)
                val_order = float(cell_order.value or 0)
            except:
                continue
            if val_forecast > 0 and val_order == 0:
                cell_forecast.fill = red_fill
                cell_order.fill = red_fill
