import re
from openpyxl import load_workbook
import pandas as pd

def extract_detail_from_click(
    file_path: str,
    summary_sheet: str,
    raw_sheet: str,
    clicked_cell: str,
    item_col_letter: str = "B",
    header_row: int = 2,
    output_path: str = None
) -> str:
    """
    根据点击的主计划单元格，从原始表中提取匹配数据，并生成新的工作表。

    参数:
        file_path: 主计划 Excel 文件路径
        summary_sheet: 主计划 sheet 名称
        raw_sheet: 原始数据 sheet 名称（如“销售明细”、“未交订单”等）
        clicked_cell: 用户点击的单元格地址（如 "F5"）
        item_col_letter: 主计划表中“品名”列（默认是 B）
        header_row: 主计划表头所在行（默认是第2行）
        output_path: 可选，另存为路径（默认覆盖原文件）

    返回:
        新 sheet 名称或错误信息
    """

    wb = load_workbook(file_path)
    ws_summary = wb[summary_sheet]
    ws_raw = wb[raw_sheet]

    row = ws_summary[clicked_cell].row
    col = ws_summary[clicked_cell].column

    item = ws_summary[f"{item_col_letter}{row}"].value
    month_field = ws_summary.cell(header_row, col).value

    if not item or not month_field:
        return f"❌ 无效点击位置，未获取品名或字段"

    match = re.match(r"(\\d{4})[-年]?(\\d{1,2})[^\\d]*(预测|订单|销售)", str(month_field))
    if not match:
        return f"❌ 无法解析字段格式：{month_field}"

    year, month, category = match.groups()
    month = month.zfill(2)
    target_year_month = f"{year}-{month}"

    # 原始表转为 DataFrame
    data = list(ws_raw.values)
    columns = data[0]
    df_raw = pd.DataFrame(data[1:], columns=columns)

    if "品名" not in df_raw.columns or "日期" not in df_raw.columns:
        return f"❌ 原始表缺少 '品名' 或 '日期' 字段"

    df_raw = df_raw[df_raw["品名"] == item].copy()
    df_raw["日期"] = pd.to_datetime(df_raw["日期"], errors="coerce")
    df_raw = df_raw[df_raw["日期"].dt.strftime("%Y-%m") == target_year_month]

    if df_raw.empty:
        return f"⚠️ 未找到匹配数据：{item} - {target_year_month}"

    # 写入新工作表
    new_sheet = f"原始-{item}-{target_year_month}"
    if new_sheet in wb.sheetnames:
        del wb[new_sheet]
    ws_new = wb.create_sheet(title=new_sheet)

    for r_idx, row in enumerate([df_raw.columns.tolist()] + df_raw.values.tolist(), 1):
        for c_idx, val in enumerate(row, 1):
            ws_new.cell(row=r_idx, column=c_idx, value=val)

    # 保存文件
    output_path = output_path or file_path
    wb.save(output_path)

    return f"✅ 已写入：{new_sheet}（保存为：{output_path}）"
